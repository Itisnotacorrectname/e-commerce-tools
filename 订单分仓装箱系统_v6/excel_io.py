"""
Excel 模板生成 + 结果报告输出
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import datetime


# ── 颜色常量 ──
CLR_HEADER_BG   = "1F4E79"   # 深蓝
CLR_HEADER_FG   = "FFFFFF"
CLR_SUBHDR_BG   = "2E75B6"
CLR_SECTION_BG  = "D6E4F0"
CLR_WARN_BG     = "FFE699"
CLR_OK_BG       = "C6EFCE"
CLR_BAD_BG      = "FFC7CE"
CLR_ALT_ROW     = "F2F7FB"


def _header_style(cell, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, size=11, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border(cell, style="thin"):
    s = Side(style=style)
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_input_template(filepath: str):
    """创建带有说明的输入模板 Excel"""
    wb = Workbook()

    # ── Sheet 1: 使用说明 ──
    ws_guide = wb.active
    ws_guide.title = "📋使用说明"
    _write_guide(ws_guide)

    # ── Sheet 2: SKU基础信息 ──
    ws_sku = wb.create_sheet("SKU信息")
    _write_sku_sheet(ws_sku)

    # ── Sheet 3: 订单数量 ──
    ws_order = wb.create_sheet("订单数量")
    _write_order_sheet(ws_order)

    # ── Sheet 4: 目标比例 ──
    ws_ratio = wb.create_sheet("目标比例")
    _write_ratio_sheet(ws_ratio)

    # ── Sheet 5: 库存在途 ──
    ws_stock = wb.create_sheet("库存在途")
    _write_stock_sheet(ws_stock)

    wb.save(filepath)


def _write_guide(ws):
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 70

    title_cell = ws['A1']
    ws.merge_cells('A1:B1')
    title_cell.value = "📦 订单分仓装箱系统 - 使用说明"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    guide_data = [
        ("", ""),
        ("📋 操作步骤", ""),
        ("步骤1", "在「SKU信息」Sheet中填写每个SKU的编号和单件体积(m³)"),
        ("步骤2", "在「订单数量」Sheet中填写本次工厂订单的每个SKU数量"),
        ("步骤3", "在「目标比例」Sheet中填写每个SKU在每个仓库的目标库存比例（每行之和须=1）"),
        ("步骤4", "在「库存在途」Sheet中填写各仓库当前在库数量、在途数量及预计到货日期"),
        ("步骤5", "保存本文件，然后运行 run.py 程序"),
        ("步骤6", "程序将生成「输出报告.xlsx」，包含分仓结果和装箱清单"),
        ("", ""),
        ("⚠️ 注意事项", ""),
        ("比例要求", "同一SKU所有仓库的目标比例之和必须等于1（例：0.4 + 0.35 + 0.25 = 1.0）"),
        ("体积单位", "请使用立方米(m³)作为体积单位"),
        ("集装箱规格", "最大容积66.0m³，最小利用率98%，每箱最多18种SKU"),
        ("日期格式", "预计到货日期格式：YYYY-MM-DD，例如 2025-08-15"),
        ("", ""),
        ("📊 输出报告说明", ""),
        ("分仓明细", "每个SKU在每个仓库的分配数量及分配前后比例对比"),
        ("装箱清单", "每个集装箱的目标仓库、装载SKU明细、利用率"),
        ("汇总统计", "集装箱总数、平均利用率、警告信息"),
    ]

    for row_idx, (key, val) in enumerate(guide_data, start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=val)
        if key in ("📋 操作步骤", "⚠️ 注意事项", "📊 输出报告说明"):
            for c in [1, 2]:
                cell = ws.cell(row=row_idx, column=c)
                cell.font = Font(name="Arial", bold=True, size=11, color=CLR_SUBHDR_BG)
                cell.fill = PatternFill("solid", start_color=CLR_SECTION_BG)
        elif key.startswith("步骤"):
            ws.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_idx, column=2).font = Font(name="Arial", size=10)


def _write_sku_sheet(ws):
    ws.title = "SKU信息"
    headers = ["SKU编号", "SKU名称（可选）", "单件体积(m³)", "备注"]
    widths = [18, 25, 18, 30]
    _write_header_row(ws, headers, widths)

    # 示例数据
    samples = [
        ["SKU-001", "示例商品A", 0.025, ""],
        ["SKU-002", "示例商品B", 0.040, ""],
        ["SKU-003", "示例商品C", 0.018, ""],
    ]
    for i, row in enumerate(samples, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10, color="808080", italic=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)

    ws.cell(row=2, column=1).comment = None
    _add_note(ws, row=1, col=3, text="请填写单件商品的体积，单位：立方米(m³)")


def _write_order_sheet(ws):
    ws.title = "订单数量"
    headers = ["SKU编号", "本次订单数量（件）", "备注"]
    widths = [18, 22, 30]
    _write_header_row(ws, headers, widths)

    samples = [
        ["SKU-001", 500, ""],
        ["SKU-002", 300, ""],
        ["SKU-003", 800, ""],
    ]
    for i, row in enumerate(samples, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10, color="808080", italic=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)


def _write_ratio_sheet(ws):
    ws.title = "目标比例"
    headers = ["SKU编号", "仓库名称", "目标比例", "备注"]
    widths = [18, 20, 15, 40]
    _write_header_row(ws, headers, widths)

    samples = [
        ["SKU-001", "美国仓", 0.40, ""],
        ["SKU-001", "欧洲仓", 0.35, ""],
        ["SKU-001", "日本仓", 0.25, "同一SKU所有仓库比例之和=1"],
        ["SKU-002", "美国仓", 0.50, ""],
        ["SKU-002", "欧洲仓", 0.30, ""],
        ["SKU-002", "日本仓", 0.20, ""],
        ["SKU-003", "美国仓", 0.45, ""],
        ["SKU-003", "欧洲仓", 0.30, ""],
        ["SKU-003", "日本仓", 0.25, ""],
    ]
    for i, row in enumerate(samples, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10, color="808080", italic=True)
            if j == 3:
                c.number_format = "0.00%"
            if i % 2 == 0:
                c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)


def _write_stock_sheet(ws):
    ws.title = "库存在途"
    headers = ["SKU编号", "仓库名称", "在库数量（件）", "在途数量（件）", "预计到货日期", "备注"]
    widths = [18, 20, 18, 18, 18, 30]
    _write_header_row(ws, headers, widths)

    samples = [
        ["SKU-001", "美国仓", 120, 80, "2025-08-10", ""],
        ["SKU-001", "欧洲仓", 90, 60, "2025-08-15", ""],
        ["SKU-001", "日本仓", 40, 30, "2025-08-20", ""],
        ["SKU-002", "美国仓", 200, 0, "", ""],
        ["SKU-002", "欧洲仓", 150, 50, "2025-08-12", ""],
        ["SKU-002", "日本仓", 80, 0, "", ""],
        ["SKU-003", "美国仓", 300, 100, "2025-08-18", ""],
        ["SKU-003", "欧洲仓", 200, 80, "2025-08-22", ""],
        ["SKU-003", "日本仓", 150, 50, "2025-08-25", ""],
    ]
    for i, row in enumerate(samples, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10, color="808080", italic=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)


def _write_header_row(ws, headers, widths):
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        _header_style(c)
        _border(c)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _add_note(ws, row, col, text):
    from openpyxl.comments import Comment
    comment = Comment(text, "System")
    ws.cell(row=row, column=col).comment = comment


# ─────────────────────────────────────────────
# 读取输入模板
# ─────────────────────────────────────────────

def read_input_template(filepath: str):
    """读取输入Excel，返回各模块数据"""
    wb = load_workbook(filepath, data_only=True)

    # SKU体积
    ws_sku = wb["SKU信息"]
    sku_volumes = {}
    sku_names = {}
    for row in ws_sku.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            sku_id = str(row[0]).strip()
            sku_volumes[sku_id] = float(row[2])
            sku_names[sku_id] = str(row[1]).strip() if row[1] else sku_id

    # 订单数量
    ws_order = wb["订单数量"]
    order_totals = {}
    for row in ws_order.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            sku_id = str(row[0]).strip()
            order_totals[sku_id] = int(row[1])

    # 目标比例
    ws_ratio = wb["目标比例"]
    target_ratios = {}
    for row in ws_ratio.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2] is not None:
            sku_id = str(row[0]).strip()
            wh = str(row[1]).strip()
            ratio = float(row[2])
            if sku_id not in target_ratios:
                target_ratios[sku_id] = {}
            target_ratios[sku_id][wh] = ratio

    # 库存在途
    from allocator import WarehouseStock
    stocks = []
    ws_stock = wb["库存在途"]
    for row in ws_stock.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            stocks.append(WarehouseStock(
                sku_id=str(row[0]).strip(),
                warehouse=str(row[1]).strip(),
                on_hand=int(row[2]) if row[2] else 0,
                in_transit=int(row[3]) if row[3] else 0,
                transit_eta=str(row[4]) if row[4] else "",
            ))

    return sku_volumes, sku_names, order_totals, target_ratios, stocks


# ─────────────────────────────────────────────
# 写出报告
# ─────────────────────────────────────────────

def write_output_report(
    filepath: str,
    allocation_results,
    containers,
    warnings: list,
    sku_names: dict = None,
):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "📊汇总"

    ws_alloc = wb.create_sheet("分仓明细")
    ws_container = wb.create_sheet("装箱清单")
    ws_warn = wb.create_sheet("⚠️警告")

    _write_summary_sheet(ws_summary, allocation_results, containers, warnings)
    _write_allocation_sheet(ws_alloc, allocation_results, sku_names)
    _write_container_sheet(ws_container, containers, sku_names)
    _write_warnings_sheet(ws_warn, warnings)

    wb.save(filepath)



def _write_summary_sheet(ws, alloc_results, containers, warnings):
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22

    # 标题
    ws.merge_cells('A1:C1')
    ws['A1'] = f"📦 订单分仓装箱结果报告  |  生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A1'].font = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # 集装箱统计
    total_containers = len(containers)
    low_util = [c for c in containers if c.is_low_utilization]
    avg_util = sum(c.utilization for c in containers) / total_containers if containers else 0
    total_vol = sum(c.used_volume for c in containers)

    stats = [
        ("", ""),
        ("📦 集装箱统计", ""),
        ("集装箱总数", f"{total_containers} 个"),
        ("平均利用率", f"{avg_util:.1%}"),
        ("总装载体积", f"{total_vol:.2f} m³"),
        ("低利用率箱数 (<98%)", f"{len(low_util)} 个"),
        ("", ""),
        ("📋 分仓统计", ""),
        ("涉及SKU数量", f"{len(set(r.sku_id for r in alloc_results))} 个"),
        ("涉及仓库数量", f"{len(set(r.warehouse for r in alloc_results))} 个"),
        ("总分配件数", f"{sum(r.allocated_qty for r in alloc_results):,} 件"),
        ("", ""),
        ("⚠️ 警告数量", f"{len(warnings)} 条"),
    ]

    for row_idx, (key, val) in enumerate(stats, start=2):
        c1 = ws.cell(row=row_idx, column=1, value=key)
        c2 = ws.cell(row=row_idx, column=2, value=val)
        if key in ("📦 集装箱统计", "📋 分仓统计"):
            c1.font = Font(name="Arial", bold=True, size=11, color=CLR_SUBHDR_BG)
            c1.fill = PatternFill("solid", start_color=CLR_SECTION_BG)
            c2.fill = PatternFill("solid", start_color=CLR_SECTION_BG)
        else:
            c1.font = Font(name="Arial", size=10, bold=True)
            c2.font = Font(name="Arial", size=10)

    # 各仓库装箱汇总表
    row = len(stats) + 4
    ws.cell(row=row, column=1, value="各仓库装箱明细").font = Font(name="Arial", bold=True, size=11, color=CLR_SUBHDR_BG)
    row += 1
    hdrs = ["仓库", "集装箱数量", "平均利用率", "总装载体积(m³)"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=row, column=j, value=h)
        _header_style(c, bg=CLR_SUBHDR_BG)
        _border(c)

    from collections import defaultdict
    wh_containers = defaultdict(list)
    for ct in containers:
        wh_containers[ct.warehouse].append(ct)

    for i, wh in enumerate(sorted(wh_containers.keys()), start=1):
        cts = wh_containers[wh]
        avg = sum(c.utilization for c in cts) / len(cts)
        total = sum(c.used_volume for c in cts)
        r = row + i
        ws.cell(row=r, column=1, value=wh).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=len(cts)).font = Font(name="Arial", size=10)
        c_util = ws.cell(row=r, column=3, value=avg)
        c_util.number_format = "0.0%"
        c_util.font = Font(name="Arial", size=10)
        if avg < 0.98:
            c_util.fill = PatternFill("solid", start_color=CLR_WARN_BG)
        else:
            c_util.fill = PatternFill("solid", start_color=CLR_OK_BG)
        ws.cell(row=r, column=4, value=round(total, 3)).font = Font(name="Arial", size=10)
        if i % 2 == 0:
            for j in [1, 2, 4]:
                ws.cell(row=r, column=j).fill = PatternFill("solid", start_color=CLR_ALT_ROW)


def _write_allocation_sheet(ws, results, sku_names):
    headers = [
        "SKU编号", "SKU名称", "仓库",
        "订单总量", "已有量\n(在库+在途)", "目标比例",
        "目标量", "本次分配量", "分配后实际比例", "偏差"
    ]
    widths = [14, 18, 14, 12, 18, 12, 12, 14, 18, 10]
    _write_header_row(ws, headers, widths)

    results_sorted = sorted(results, key=lambda r: (r.sku_id, r.warehouse))
    for i, r in enumerate(results_sorted, start=2):
        name = (sku_names or {}).get(r.sku_id, "")
        deviation = r.final_ratio - r.target_ratio
        row_data = [
            r.sku_id, name, r.warehouse,
            r.order_qty, r.already_qty, r.target_ratio,
            r.target_qty, r.allocated_qty, r.final_ratio, deviation
        ]
        for j, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = PatternFill("solid", start_color=CLR_ALT_ROW)

        # 格式化比例列
        ws.cell(row=i, column=6).number_format = "0.00%"
        ws.cell(row=i, column=9).number_format = "0.00%"
        ws.cell(row=i, column=10).number_format = "+0.00%;-0.00%"

        # 偏差着色
        dev_cell = ws.cell(row=i, column=10)
        if abs(deviation) > 0.05:
            dev_cell.fill = PatternFill("solid", start_color=CLR_BAD_BG)
        elif abs(deviation) > 0.02:
            dev_cell.fill = PatternFill("solid", start_color=CLR_WARN_BG)
        else:
            dev_cell.fill = PatternFill("solid", start_color=CLR_OK_BG)

        # 分配量为0高亮
        if r.allocated_qty == 0:
            ws.cell(row=i, column=8).fill = PatternFill("solid", start_color=CLR_WARN_BG)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(results_sorted)+1}"


def _write_container_sheet(ws, containers, sku_names):
    headers = [
        "集装箱编号", "目标仓库", "SKU编号", "SKU名称",
        "数量(件)", "该SKU体积(m³)",
        "箱内SKU种数", "已用体积(m³)", "利用率", "状态"
    ]
    widths = [14, 14, 14, 18, 12, 18, 14, 16, 12, 12]
    _write_header_row(ws, headers, widths)

    row = 2
    for ct in containers:
        util = ct.utilization
        status = "✅正常" if not ct.is_low_utilization else "⚠️低利用率"
        sku_count = ct.sku_count

        # 按SKU分组展示
        sku_summary = {}
        for sku_id, qty, vol in ct.items:
            if sku_id not in sku_summary:
                sku_summary[sku_id] = [0, 0.0]
            sku_summary[sku_id][0] += qty
            sku_summary[sku_id][1] += vol

        for idx, (sku_id, (qty, vol)) in enumerate(sku_summary.items()):
            name = (sku_names or {}).get(sku_id, "")
            row_data = [
                ct.container_id if idx == 0 else "",
                ct.warehouse if idx == 0 else "",
                sku_id, name, qty, round(vol, 4),
                sku_count if idx == 0 else "",
                round(ct.used_volume, 3) if idx == 0 else "",
                util if idx == 0 else "",
                status if idx == 0 else "",
            ]
            for j, val in enumerate(row_data, start=1):
                c = ws.cell(row=row, column=j, value=val)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center")
                bg = CLR_BAD_BG if ct.is_low_utilization else (CLR_ALT_ROW if row % 2 == 0 else "FFFFFF")
                c.fill = PatternFill("solid", start_color=bg)

            if idx == 0:
                util_cell = ws.cell(row=row, column=9)
                util_cell.number_format = "0.0%"
                util_cell.font = Font(name="Arial", size=10, bold=True)
                if ct.is_low_utilization:
                    util_cell.fill = PatternFill("solid", start_color=CLR_BAD_BG)
                else:
                    util_cell.fill = PatternFill("solid", start_color=CLR_OK_BG)

            row += 1

        # 分隔线
        for j in range(1, 11):
            c = ws.cell(row=row, column=j)
            c.border = Border(bottom=Side(style="medium"))
        row += 1

    ws.freeze_panes = "A2"


def _write_warnings_sheet(ws, warnings):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80

    ws.merge_cells('A1:B1')
    ws['A1'] = "⚠️ 运行警告与提示"
    ws['A1'].font = Font(name="Arial", bold=True, size=12, color="CC0000")
    ws['A1'].alignment = Alignment(horizontal="center")

    if not warnings:
        ws.cell(row=2, column=1, value="✅").font = Font(size=12)
        ws.cell(row=2, column=2, value="无警告，所有计算正常完成").font = Font(name="Arial", size=11, color="00AA00")
        return

    for i, w in enumerate(warnings, start=2):
        ws.cell(row=i, column=1, value=i-1).font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=i, column=2, value=w)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=CLR_WARN_BG)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 20
